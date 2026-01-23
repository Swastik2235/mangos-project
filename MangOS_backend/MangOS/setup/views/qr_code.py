from django.shortcuts import get_object_or_404
from django.http import FileResponse,JsonResponse
from setup.models import MachineQRCode,MachineMaster
from rest_framework import viewsets

import pandas as pd
from rest_framework.response import Response
from rest_framework import status
from rest_framework import viewsets
from django.core.files.storage import default_storage
from setup.models import BOMSheet, JobCard, ClientDetails, InventoryItem,ItemMaster
from setup.serializers.job_card_serializers import CreateJobCardSerializer, GetJobCardSerializer
from django.http import HttpResponse,JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os
from django.conf import settings



class QRCodeViewset(viewsets.ModelViewSet):

    def get_machine_qr_code(self,request):
        
        machine_id = request.GET.get('machine_id')
        qr_obj = get_object_or_404(MachineQRCode, machine_id=machine_id)
    
        try:
            # This uses the relative path stored in the ImageField
            relative_path = qr_obj.qr_code.name  # e.g., "qr_codes/31_qr.png"
            file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
            print('file_path:', file_path)
    
            if not os.path.exists(file_path):
                # Fallback to placeholder
                placeholder_path = os.path.join(settings.MEDIA_ROOT, 'qr_codes', 'placeholder.png')
                if os.path.exists(placeholder_path):
                    return FileResponse(open(placeholder_path, 'rb'), content_type='image/png')
                return JsonResponse({"detail": "QR code file not found on disk."}, status=404)
        except ValueError:
            return JsonResponse({"detail": "QR code not found."}, status=404)
    
        return FileResponse(open(file_path, 'rb'), content_type='image/png')


      
        
        




def _safe_float(value, default=0.0):
    try:
        return float(value) if value else default
    except (ValueError, TypeError):
        return default

def _safe_int(value, default=0):
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default

class CreateJobCardViewSet(viewsets.ModelViewSet):
    


    def _add_outside_border(self, ws, start_row, end_row):
        """Add thick border only around the outside of the job card section"""
        thick = Side(style='thick')
        thin = Side(style='thin')
        
        
        # Top border (thick on top, thin on other sides)
        for col in range(1, 20):  # Columns A to S
            cell = ws.cell(row=start_row, column=col)
            if cell.border is None:
                cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(top=thick, 
                                   left=cell.border.left or thin,
                                   right=cell.border.right or thin,
                                   bottom=cell.border.bottom or thin)
        
        # Bottom border
        for col in range(1, 20):
            cell = ws.cell(row=end_row, column=col)
            if cell.border is None:
                cell.border = Border(bottom=thick, left=thin, right=thin, top=thin)
            else:
                cell.border = Border(bottom=thick,
                                    left=cell.border.left or thin,
                                    right=cell.border.right or thin,
                                    top=cell.border.top or thin)
        
        # Left border (first column)
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.border is None:
                cell.border = Border(left=thick, top=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(left=thick,
                                   top=cell.border.top or thin,
                                   right=cell.border.right or thin,
                                   bottom=cell.border.bottom or thin)
        
        # Right border (last column)
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=19)
            if cell.border is None:
                cell.border = Border(right=thick, top=thin, left=thin, bottom=thin)
            else:
                cell.border = Border(right=thick,
                                   top=cell.border.top or thin,
                                   left=cell.border.left or thin,
                                   bottom=cell.border.bottom or thin)

    def _create_job_card_header(self, ws, start_row, job):
        """Create the complete header section for each job card"""
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

        # Company info
        ws.merge_cells(f'A{start_row}')
        ws[f'A{start_row}'] = 'ERA T & D Ltd.,'
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        
        ws.merge_cells(f'A{start_row+1}')
        ws[f'A{start_row+1}'] = 'Umred.'

        # Document title
        ws.merge_cells(f'K{start_row}')
        ws[f'K{start_row}'] = 'JOB CARD'
        ws[f'K{start_row}'].font = Font(bold=True)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'K{start_row+1}')
        ws[f'K{start_row+1}'] = 'CUM'
        ws[f'K{start_row+1}'].font = Font(size=8,bold=True)
        ws[f'K{start_row+1}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'J{start_row+2}:L{start_row+2}')
        ws[f'J{start_row+2}'] = 'STEEL REQUISITION SLIP'
        ws[f'J{start_row+2}'].font = Font(size=8,bold=True)
        ws[f'J{start_row+2}'].alignment = Alignment(horizontal='center')

        # Format numbers
        ws.merge_cells(f'P{start_row+3}:S{start_row+3}')
        ws[f'P{start_row+3}'] = 'Format No:-'
        ws[f'P{start_row+3}'].font = Font(bold=True)
        ws[f'P{start_row+3}'].border = thin_border

        # Job details
        details = [
            (f"A{start_row+4}:B{start_row+4}", f"Job Order No         :  {job.Job_order_no if job.Job_order_no else 'N/A'}"),
            (f"A{start_row+5}:C{start_row+5}", f"Client                       :  {job.client_details_id.client_name if job.client_details_id else 'N/A'}"),
            (f"A{start_row+6}:D{start_row+6}", f"Project                    :  {job.project_name if job.project_name else 'N/A'}"),
            (f"A{start_row+7}:C{start_row+7}", f"Type  of  Tower     :  {job.type_of_tower if job.type_of_tower else 'N/A'}"),
            (f"A{start_row+8}:D{start_row+8}", "Please  fabricate  the  following  Mark  Nos. against with Mass Production."),
            

        ]

        for cell_range, text in details:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = text
            cell.font = Font(size=11,bold=True)
            cell.alignment = Alignment(vertical="center")

        # Job Details (Right Side)
        ws[f'O{start_row+4}'] = 'Released Date :-'
        # ws[f'P{start_row+4}'] = job.release_date.strftime('%d-%b-%y') if job.release_date else 'N/A'

        ws[f'O{start_row+6}'] = 'JC No. :'
        # ws[f'P{start_row+5}'] = f"JC / 162 / {job.Job_order_no.split('-')[-1]}" if job.Job_order_no else 'N/A'

        ws[f'O{start_row+7}'] = 'Speci. :'
        # ws[f'P{start_row+6}'] = job.specification if hasattr(job, 'specification') else ''

        ws[f'O{start_row+8}'] = 'LOT :'
        # ws[f'P{start_row+7}'] = job.lot_no if hasattr(job, 'lot_no') else '1'

        

        return start_row + 9


        

    def _create_table_headers(self, ws, row):
        """Create table headers for each job card section with full borders"""

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers_row1 = [
            'MARK', '', 'SECTION', '', 'LENGTH', 'WIDTH', 'UNIT', 'WT/PCS', '', '', 
            'TOTAL', 'TOTAL', 'Operations', '', '', '', '', '', 'REMARK'
        ]
        headers_row2 = [
            'NO.', '', '( mm )', '', '( mm )', '( mm )', 'WT.', '', '', '', 
            'QTY.', 'WEIGHT', '', 'N', 'B', 'HAB', 'HC', 'D', ''
        ]

        num_columns = len(headers_row1)

        # Write first header row with border (even if empty)
        for col_num in range(1, num_columns + 1):
            value = headers_row1[col_num - 1]
            cell = ws.cell(row=row + 1, column=col_num, value=value if value else None)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Write second header row with border (even if empty)
        for col_num in range(1, num_columns + 1):
            value = headers_row2[col_num - 1]
            cell = ws.cell(row=row + 2, column=col_num, value=value if value else None)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Merge necessary cells
        merges = [
            f'A{row+1}:B{row+1}',  # MARK
            f'A{row+2}:B{row+2}',  # MARK NO.
            f'C{row+1}:D{row+1}',  # SECTION
            f'C{row+2}:D{row+2}',  # (mm)
            f'E{row+1}:E{row+2}',  # LENGTH
            f'F{row+1}:F{row+2}',  # WIDTH
            f'G{row+1}:G{row+2}',  # UNIT WT.
            f'H{row+1}:J{row+1}',  # WT/PCS
            f'H{row+2}:J{row+2}',  # WT/PCS cells (3 columns wide)
            f'K{row+1}:K{row+2}',  # TOTAL QTY.
            f'L{row+1}:L{row+2}',  # TOTAL WEIGHT
            f'M{row+1}:R{row+1}',  # Operations
            f'M{row+2}:M{row+2}',  # N
            f'N{row+2}:N{row+2}',  # B
            f'O{row+2}:O{row+2}',  # HAB
            f'P{row+2}:P{row+2}',  # HC
            f'Q{row+2}:Q{row+2}',  # D
            f'S{row+1}:S{row+2}'   # REMARK
        ]
        for merge_range in merges:
            ws.merge_cells(merge_range)

        return row + 3

    def _create_job_entries(self, ws, start_row, section_jobs):
        """Create job entries with perfect cell borders and design"""
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate adjusted row range
        first_row = max(1, start_row - 1)  # Start one row before (minimum row 1)
        last_row = start_row + len(section_jobs) * 2 + 1  # End one row before original
        
        # Apply perfect borders to all relevant cells
        for row in range(first_row, last_row + 1):
            for col in range(1, 20):  # Columns A-S
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                # Clear values in border-only rows (empty spacer rows)
                if row not in range(start_row, last_row, 2):
                    cell.value = ""

        current_row = start_row

        # Create job entries with existing design
        for job in section_jobs:
            row_data = [
                job.mark_no or "", "",                  # A-B: MARK NO
                job.section or "", "",                  # C-D: SECTION
                job.length or "",                       # E: LENGTH
                job.width or "",                        # F: WIDTH
                job.unit or "",                         # G: UNIT
                job.wt_and_pcs or "", "", "",           # H-J: WT/PCS + Empty
                job.total_quantity or "",               # K: TOTAL QTY
                job.total_weight or "",                 # L: TOTAL WEIGHT
                *[""] * 7                               # M-S: (empty columns)
            ]

            # Write data with existing formatting
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                if col_num in [1,3,4,5,6,7, 8, 11, 12]:  # Right-align numeric columns
                    cell.alignment = Alignment(horizontal='center')

            current_row += 2  # Maintain double spacing

        # Add section total with existing design
        section_qty = sum(_safe_int(job.total_quantity) for job in section_jobs)
        section_wt = sum(_safe_float(job.total_weight) for job in section_jobs)

        ws.merge_cells(f'A{current_row}:G{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value='Section Total')
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=current_row, column=11, value=section_qty).font = Font(bold=True)
        ws.cell(row=current_row, column=12, value=round(section_wt, 2)).font = Font(bold=True)
        
        # Right-align total values
        for col in [11, 12]:
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        return current_row + 2 # Return one row earlier than original



    def generate_section(self, ws, start_row, section_jobs):
        """Helper to create headers + jobs"""
        header_end_row = self._create_table_headers(ws, start_row)
        job_start_row = header_end_row + 1  # Leave 1 empty row after headers
        return self._create_job_entries(ws, job_start_row, section_jobs)


    def _create_allocation_section(self, ws, start_row, section_jobs):
        # Get item names from InventoryItem and lengths from ItemMaster
        inventory_items = InventoryItem.objects.values_list("id", "item_name")
        
        # Assuming ItemMaster has a field named 'name' or similar that corresponds to item_name
        # Replace 'name' with the actual field name in ItemMaster that matches InventoryItem's item_name
        item_lengths = {item.name: item.length for item in ItemMaster.objects.all()}  # Change 'name' to the correct field
    
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
    
        headers = ['', '', '', 'SR.NO.', 'Alloca.-1', '', 'Alloca.-2', '', '', '', 'Alloca.-3', '', 
                 'Alloca.-4', '', 'Consu length', 'Length', 'Pcs', 'C.L.', 'QTY']
    
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
        merges = [
            f'D{start_row}',        # SR.NO
            f'E{start_row}:F{start_row}',  # Alloca.-1
            f'G{start_row}:H{start_row}',  # Alloca.-2
            f'K{start_row}:L{start_row}',  # Alloca.-3
            f'M{start_row}:N{start_row}',  # Alloca.-4
            f'O{start_row}',        # Consu length
            f'P{start_row}',        # Length
            f'Q{start_row}',        # Pcs
            f'R{start_row}',        # C.L.
            f'S{start_row}'         # QTY
        ]
        for m in merges:
            ws.merge_cells(m)
    
        current_row = start_row + 1
        allocation_data = []
    
        for i, job in enumerate(section_jobs[:2]):
            try:
                job_length = float(job.length)
            except (TypeError, ValueError):
                job_length = None
    
            # Find matching length by checking item_name in InventoryItem and length in ItemMaster
            matching_length = None
            for _, item_name in inventory_items:
                if item_name == job.section:
                    # Get the corresponding field from ItemMaster that matches item_name
                    # This assumes there's a common identifier between the two models
                    item_master = ItemMaster.objects.filter(name=item_name).first()  # Change 'name' to correct field
                    if item_master:
                        try:
                            matching_length = float(item_master.length) if item_master.length else None
                        except (TypeError, ValueError):
                            matching_length = None
                    break
    
            # Calculate count, cutting loss, and consumed length
            if job_length and matching_length and job_length > 0:
                count = int(matching_length // job_length)
                cutting_loss = matching_length - (count * job_length)
                consumed_length = count * job_length
            else:
                count = ''
                cutting_loss = ''
                consumed_length = ''
    
            try:
                pcs = round(float(job.total_quantity) / count, 2) if count else ""
            except (TypeError, ValueError, ZeroDivisionError):
                pcs = ""    
    
            allocation_data.append([
                '', '', '',                    # A-C
                i + 1,                        # D - SR.NO
                job.length or "",             # E - Alloca.-1
                count,                        # F - Count (calculated)
                '', '', '', '', '', '', '', '', # G-L
                consumed_length,              # M - Consu length
                matching_length,              # N - Length (from ItemMaster)
                pcs,
                cutting_loss,                 # P - C.L.
                job.total_quantity or ""      # Q - QTY
            ])
    
        # Write allocation data with spacing
        for row_data in allocation_data:
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [4,5,6,13,14,15,16,17,18,19]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center')
            current_row += 2  # Spacing
    
        return current_row

        

    def _create_embossing_section(self, ws, start_row,section_jobs):
        """Create the embossing and weight summary section with dynamic data"""

        # Get item names from InventoryItem and lengths from ItemMaster
        inventory_items = InventoryItem.objects.values_list("id", "item_name")
        item_lengths = {item.name: item.length for item in ItemMaster.objects.all()}
    
        for i, job in enumerate(section_jobs[:2]):
            try:
                job_length = float(job.length)
            except (TypeError, ValueError):
                job_length = None
    
            # Find matching length by checking item_name in both models
            matching_length = None
            for _, item_name in inventory_items:
                if item_name == job.section:
                    try:
                        matching_length = float(item_lengths.get(item_name, 0))
                    except (TypeError, ValueError):
                        matching_length = None
                    break
    
        thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    
        # Embossing header
        ws.merge_cells(f'A{start_row}:M{start_row}')
        ws[f'A{start_row}'] = 'EMBOSSING DETAILS'
        ws[f'A{start_row}'].font = Font(bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{start_row}'].border = thin_border
    
        # Embossing data headers
        headers = ['EMBOSSING', 'SIN NO', 'SECTION', '', 'Length', 'pcs', 'Unit wt', 'wt in kgs', '', '', 'C.L', 'Pcs', 'Wt in kgs']
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
    
        # Generate embossing data from section_jobs
        embossing_data = []
        for job in section_jobs:
            try:
                unit_wt = float(job.wt_and_pcs) if job.wt_and_pcs else 0
                total_qty = float(job.total_quantity) if job.total_quantity else 0
                wt_in_kgs = round(unit_wt * total_qty, 2)
            except (ValueError, TypeError):
                unit_wt = 0
                wt_in_kgs = 0
    
            # Get length for this job's section from ItemMaster
            section_length = None
            for _, item_name in inventory_items:
                if item_name == job.section:
                    section_length = item_lengths.get(item_name)
                    break
    
            embossing_data.append([
                'MD PG MS21',
                'SIN-36 21-22',
                job.section or "",
                "",
                float(section_length) if section_length else "",
                job.total_quantity or "",
                unit_wt,
                wt_in_kgs,
                "", "",
                "",
                "",
                "Kgs"
            ])
    
        # Write embossing data
        for row_num, row_data in enumerate(embossing_data, start_row+2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [5, 6, 7, 8, 12]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center')
    
        # Calculate weight summary from the actual data
        total_planning_wt = sum(row_data[7] for row_data in embossing_data if isinstance(row_data[7], (int, float)))
        total_raw_material_wt = total_planning_wt * 1.01
        wastage_percent = 1.05
        wastage_wt = round(total_planning_wt * (wastage_percent / 100), 2)
    
        # Weight summary
        summary_start = start_row + len(embossing_data) -1
        summary = [
            ('Planning Wt.                      :', total_planning_wt, 'Kgs'),
            ('Raw Material Wt.                  :', round(total_raw_material_wt, 2), 'Kgs'),
            ('SRN Wt                            :', '0', 'Kgs'),
            ('Wastage                           :', wastage_percent, '%'),
            ('Wastage Wt.                       :', wastage_wt, 'Kgs')
        ]
    
        for i, (label, value, unit) in enumerate(summary):
            ws[f'O{summary_start + i}'] = label
            ws[f'O{summary_start + i}'].font = Font(bold=True)
            ws[f'O{summary_start + i}'].alignment = Alignment(horizontal='right')
            
            ws[f'R{summary_start + i}'] = value
            ws[f'R{summary_start + i}'].alignment = Alignment(horizontal='right')
            
            ws[f'S{summary_start + i}'] = unit
            ws[f'S{summary_start + i}'].alignment = Alignment(horizontal='left')
    
        total_pcs = sum(_safe_int(job.total_quantity) for job in section_jobs)
        total_wt = sum(_safe_float(row[6]) for row in embossing_data if isinstance(row[6], (int, float)))
    
        # Add total row
        total_row = start_row + len(embossing_data) + 4
        ws.merge_cells(f'C{total_row}')
        ws[f'C{total_row}'] = 'Total:'
        ws[f'C{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].alignment = Alignment(horizontal='center')
        
        ws[f'F{total_row}'] = total_pcs
        ws[f'F{total_row}'].font = Font(bold=True)
        ws[f'F{total_row}'].alignment = Alignment(horizontal='center')
        
        ws[f'H{total_row}'] = round(total_wt, 2)
        ws[f'H{total_row}'].font = Font(bold=True)
        ws[f'H{total_row}'].alignment = Alignment(horizontal='center')
    
        # Add JC number
        ws.merge_cells(f'O{total_row}:S{total_row}')
        ws[f'O{total_row}'] = f"JC / 162 / {section_jobs[0].Job_order_no.split('-')[-1]}" if section_jobs and section_jobs[0].Job_order_no else 'JC / 162 / 000'
        ws[f'O{total_row}'].font = Font(bold=True)
        ws[f'O{total_row}'].alignment = Alignment(horizontal='center')
    
        return summary_start + len(summary) + 1

        

    def get_job_card_by_qr(self, request):
        try:
            # Get and group job cards by section
            job_cards = JobCard.objects.all().order_by('section')
            if not job_cards:
                return Response({"error": "No job cards found"}, status=status.HTTP_404_NOT_FOUND)

            # Group by section
            section_groups = defaultdict(list)
            for job in job_cards:
                section_groups[job.section].append(job)

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Set column widths
            column_widths = {
                'A': 15, 'B': 10, 'C': 28, 'D': 15, 'E': 10, 'F': 10, 'G': 10,
                'H': 10, 'I': 5, 'J': 5, 'K': 18, 'L': 10, 'M': 10, 'N': 2,
                'O': 28, 'P': 8, 'Q': 5, 'R': 5, 'S': 15
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            current_row = 2  # Start at first row

            # Create separate job card for each section
            for section, jobs in section_groups.items():
                start_job_card_row = current_row
                
                # Create complete header for this job card
                current_row = self._create_job_card_header(ws, current_row, jobs[0])
                
                # Create table headers
                # current_row = self._create_table_headers(ws, current_row)
                
                # Create job entries for this section
                # current_row = self._create_job_entries(ws, current_row, jobs)

                current_row = self.generate_section(ws, current_row, jobs)
                
                # Add allocation section
                current_row = self._create_allocation_section(ws, current_row, jobs)
                
                # Add embossing and weight summary
                current_row = self._create_embossing_section(ws, current_row,jobs)
                
                # Add thick border only around the outside of the job card
                self._add_outside_border(ws, start_job_card_row, current_row - 1)
                
                # Add spacing between job cards
                current_row += 2

            save_dir = os.path.join(settings.MEDIA_ROOT, "job_cards")
            os.makedirs(save_dir, exist_ok=True)

            filename = "job_cards.xlsx"
            full_path = os.path.join(save_dir, filename)
            wb.save(full_path)  # 💾 Save the file on disk

            # 📤 Create response for download
            with open(full_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            
            
            
            
            
#def get_job_card_by_qr(self,request):
       # machine_id = request.GET.get('machine_id')
       # machine = get_object_or_404(MachineMaster, id=machine_id)

        # You can serialize this or return job card data
       # return JsonResponse({
       #     "machine_name": machine.machine_name,
       #     "machine_type": machine.machine_type,
       #     "amc_type": machine.amc_type,
       #     "vendor": machine.vendor,
       #     "created_at": machine.created_at,
       # })               
            
            
        