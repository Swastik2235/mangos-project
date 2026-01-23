import pandas as pd
from rest_framework.response import Response
from rest_framework import status
from rest_framework import viewsets
from django.core.files.storage import default_storage
from setup.models import BOMSheet, JobCard, ClientDetails, InventoryItem,Project,JobCardReport
from setup.serializers.job_card_serializers import CreateJobCardSerializer, GetJobCardSerializer,JobCardSerializer,JobCardReportSerializer,JobCardMinimalSerializer
from django.http import HttpResponse,JsonResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os
from django.conf import settings
import requests
import tempfile
from decimal import Decimal, InvalidOperation
from decimal import Decimal
from django.utils import timezone



def _safe_float(value, default=0.0):
    try:
        return float(value) if value else default
    except (ValueError, TypeError):
        return default

def _safe_int(value, default=0):
    try:
        return int(value) if value else default
    except (ValueError, TypeError):
        return default

class CreateJobCardViewSet(viewsets.ModelViewSet):
    serializer_class = GetJobCardSerializer


    def _add_outside_border(self, ws, start_row, end_row):
        """Add thick border only around the outside of the job card section"""
        thick = Side(style='thick')
        thin = Side(style='thin')
        
        
        # Top border (thick on top, thin on other sides)
        for col in range(1, 20):  # Columns A to S
            cell = ws.cell(row=start_row, column=col)
            if cell.border is None:
                cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(top=thick, 
                                   left=cell.border.left or thin,
                                   right=cell.border.right or thin,
                                   bottom=cell.border.bottom or thin)
        
        # Bottom border
        for col in range(1, 20):
            cell = ws.cell(row=end_row, column=col)
            if cell.border is None:
                cell.border = Border(bottom=thick, left=thin, right=thin, top=thin)
            else:
                cell.border = Border(bottom=thick,
                                    left=cell.border.left or thin,
                                    right=cell.border.right or thin,
                                    top=cell.border.top or thin)
        
        # Left border (first column)
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.border is None:
                cell.border = Border(left=thick, top=thin, right=thin, bottom=thin)
            else:
                cell.border = Border(left=thick,
                                   top=cell.border.top or thin,
                                   right=cell.border.right or thin,
                                   bottom=cell.border.bottom or thin)
        
        # Right border (last column)
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=19)
            if cell.border is None:
                cell.border = Border(right=thick, top=thin, left=thin, bottom=thin)
            else:
                cell.border = Border(right=thick,
                                   top=cell.border.top or thin,
                                   left=cell.border.left or thin,
                                   bottom=cell.border.bottom or thin)

    def _create_job_card_header(self, ws, start_row, job):
        """Create the complete header section for each job card"""
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

        # Company info
        ws.merge_cells(f'A{start_row}')
        ws[f'A{start_row}'] = 'ERA T & D Ltd.,'
        ws[f'A{start_row}'].font = Font(size=12, bold=True)
        
        ws.merge_cells(f'A{start_row+1}')
        ws[f'A{start_row+1}'] = 'Umred.'

        # Document title
        ws.merge_cells(f'K{start_row}')
        ws[f'K{start_row}'] = 'JOB CARD'
        ws[f'K{start_row}'].font = Font(bold=True)
        ws[f'K{start_row}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'K{start_row+1}')
        ws[f'K{start_row+1}'] = 'CUM'
        ws[f'K{start_row+1}'].font = Font(size=8,bold=True)
        ws[f'K{start_row+1}'].alignment = Alignment(horizontal='center')

        ws.merge_cells(f'J{start_row+2}:L{start_row+2}')
        ws[f'J{start_row+2}'] = 'STEEL REQUISITION SLIP'
        ws[f'J{start_row+2}'].font = Font(size=8,bold=True)
        ws[f'J{start_row+2}'].alignment = Alignment(horizontal='center')

        # Format numbers
        ws.merge_cells(f'P{start_row+3}:S{start_row+3}')
        ws[f'P{start_row+3}'] = 'Format No:-'
        ws[f'P{start_row+3}'].font = Font(bold=True)
        ws[f'P{start_row+3}'].border = thin_border

        # Job details
        details = [
            (f"A{start_row+4}:B{start_row+4}", f"Job Order No         :  {job.Job_order_no if job.Job_order_no else 'N/A'}"),
            #(f"A{start_row+5}:C{start_row+5}", f"Client                       :  {job.client_details_id.client_name if job.client_details_id else 'N/A'}"),
            #(f"A{start_row+6}:D{start_row+6}", f"Project                    :  {job.project_name if job.project_name else 'N/A'}"),
            (f"A{start_row+7}:C{start_row+7}", f"Type  of  Tower     :  {job.type_of_tower if job.type_of_tower else 'N/A'}"),
            (f"A{start_row+8}:D{start_row+8}", "Please  fabricate  the  following  Mark  Nos. against with Mass Production."),
            

        ]

        for cell_range, text in details:
            ws.merge_cells(cell_range)
            cell = ws[cell_range.split(":")[0]]
            cell.value = text
            cell.font = Font(size=11,bold=True)
            cell.alignment = Alignment(vertical="center")

        # Job Details (Right Side)
        ws[f'O{start_row+4}'] = 'Released Date :-'
        # ws[f'P{start_row+4}'] = job.release_date.strftime('%d-%b-%y') if job.release_date else 'N/A'

        ws[f'O{start_row+6}'] = 'JC No. :'
        # ws[f'P{start_row+5}'] = f"JC / 162 / {job.Job_order_no.split('-')[-1]}" if job.Job_order_no else 'N/A'

        ws[f'O{start_row+7}'] = 'Speci. :'
        # ws[f'P{start_row+6}'] = job.specification if hasattr(job, 'specification') else ''

        ws[f'O{start_row+8}'] = 'LOT :'
        # ws[f'P{start_row+7}'] = job.lot_no if hasattr(job, 'lot_no') else '1'

        

        return start_row + 9


        

    def _create_table_headers(self, ws, row):
        """Create table headers for each job card section with full borders"""

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers_row1 = [
            'MARK', '', 'SECTION', '', 'LENGTH', 'WIDTH', 'UNIT', 'WT/PCS', '', '', 
            'TOTAL', 'TOTAL', 'Operations', '', '', '', '', '', 'REMARK'
        ]
        headers_row2 = [
            'NO.', '', '( mm )', '', '( mm )', '( mm )', 'WT.', '', '', '', 
            'QTY.', 'WEIGHT', '', 'N', 'B', 'HAB', 'HC', 'D', ''
        ]

        num_columns = len(headers_row1)

        # Write first header row with border (even if empty)
        for col_num in range(1, num_columns + 1):
            value = headers_row1[col_num - 1]
            cell = ws.cell(row=row + 1, column=col_num, value=value if value else None)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Write second header row with border (even if empty)
        for col_num in range(1, num_columns + 1):
            value = headers_row2[col_num - 1]
            cell = ws.cell(row=row + 2, column=col_num, value=value if value else None)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Merge necessary cells
        merges = [
            f'A{row+1}:B{row+1}',  # MARK
            f'A{row+2}:B{row+2}',  # MARK NO.
            f'C{row+1}:D{row+1}',  # SECTION
            f'C{row+2}:D{row+2}',  # (mm)
            f'E{row+1}:E{row+2}',  # LENGTH
            f'F{row+1}:F{row+2}',  # WIDTH
            f'G{row+1}:G{row+2}',  # UNIT WT.
            f'H{row+1}:J{row+1}',  # WT/PCS
            f'H{row+2}:J{row+2}',  # WT/PCS cells (3 columns wide)
            f'K{row+1}:K{row+2}',  # TOTAL QTY.
            f'L{row+1}:L{row+2}',  # TOTAL WEIGHT
            f'M{row+1}:R{row+1}',  # Operations
            f'M{row+2}:M{row+2}',  # N
            f'N{row+2}:N{row+2}',  # B
            f'O{row+2}:O{row+2}',  # HAB
            f'P{row+2}:P{row+2}',  # HC
            f'Q{row+2}:Q{row+2}',  # D
            f'S{row+1}:S{row+2}'   # REMARK
        ]
        for merge_range in merges:
            ws.merge_cells(merge_range)

        return row + 3

    def _create_job_entries(self, ws, start_row, section_jobs):
        """Create job entries with perfect cell borders and design"""
        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Calculate adjusted row range
        first_row = max(1, start_row - 1)  # Start one row before (minimum row 1)
        last_row = start_row + len(section_jobs) * 2 + 1  # End one row before original
        
        # Apply perfect borders to all relevant cells
        for row in range(first_row, last_row + 1):
            for col in range(1, 20):  # Columns A-S
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                # Clear values in border-only rows (empty spacer rows)
                if row not in range(start_row, last_row, 2):
                    cell.value = ""

        current_row = start_row

        # Create job entries with existing design
        for job in section_jobs:
            row_data = [
                job.mark_no or "", "",                  # A-B: MARK NO
                job.section or "", "",                  # C-D: SECTION
                job.length or "",                       # E: LENGTH
                job.width or "",                        # F: WIDTH
                job.unit or "",                         # G: UNIT
                job.wt_and_pcs or "", "", "",           # H-J: WT/PCS + Empty
                job.total_quantity or "",               # K: TOTAL QTY
                job.total_weight or "",                 # L: TOTAL WEIGHT
                *[""] * 7                               # M-S: (empty columns)
            ]

            # Write data with existing formatting
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                if col_num in [1,3,4,5,6,7, 8, 11, 12]:  # Right-align numeric columns
                    cell.alignment = Alignment(horizontal='center')

            current_row += 2  # Maintain double spacing

        # Add section total with existing design
        section_qty = sum(_safe_int(job.total_quantity) for job in section_jobs)
        section_wt = sum(_safe_float(job.total_weight) for job in section_jobs)

        ws.merge_cells(f'A{current_row}:G{current_row}')
        total_cell = ws.cell(row=current_row, column=1, value='Section Total')
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=current_row, column=11, value=section_qty).font = Font(bold=True)
        ws.cell(row=current_row, column=12, value=round(section_wt, 2)).font = Font(bold=True)
        
        # Right-align total values
        for col in [11, 12]:
            ws.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')

        return current_row + 2 # Return one row earlier than original



    def generate_section(self, ws, start_row, section_jobs):
        """Helper to create headers + jobs"""
        header_end_row = self._create_table_headers(ws, start_row)
        job_start_row = header_end_row + 1  # Leave 1 empty row after headers
        return self._create_job_entries(ws, job_start_row, section_jobs)


    def _create_allocation_section(self, ws, start_row, section_jobs):
        inventory_data = InventoryItem.objects.select_related('item_id').values_list(
        "id", "item_id__name", "item_id__length"
        )
    
    
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))

        headers = ['', '', '', 'SR.NO.', 'Alloca.-1', '', 'Alloca.-2', '', '', '', 'Alloca.-3', '', 
                'Alloca.-4', '', 'Consu length', 'Length', 'Pcs', 'C.L.', 'QTY']

        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        merges = [
            f'D{start_row}',        # SR.NO
            f'E{start_row}:F{start_row}',  # Alloca.-1
            f'G{start_row}:H{start_row}',  # Alloca.-2
            f'K{start_row}:L{start_row}',  # Alloca.-3
            f'M{start_row}:N{start_row}',  # Alloca.-4
            f'O{start_row}',        # Consu length
            f'P{start_row}',        # Length
            f'Q{start_row}',        # Pcs
            f'R{start_row}',        # C.L.
            f'S{start_row}'         # QTY
        ]
        for m in merges:
            ws.merge_cells(m)

        current_row = start_row + 1
        allocation_data = []

        for i, job in enumerate(section_jobs[:2]):
            try:
                job_length = float(job.length)
            except (TypeError, ValueError):
                job_length = None

            # Find matching inventory length by product name == job.section
            matching_length = None
            for _, prod_name, inv_length in inventory_data:
                if prod_name == job.section:
                    try:
                        matching_length = float(inv_length)
                    except (TypeError, ValueError):
                        matching_length = None
                    break

            # Calculate count, cutting loss, and consumed length
            if job_length and matching_length and job_length > 0:
                count = int(matching_length // job_length)
                cutting_loss = matching_length - (count * job_length)
                consumed_length = count * job_length
            else:
                count = ''
                cutting_loss = ''
                consumed_length = ''


            try:
                pcs = round(float(job.total_quantity) / count, 2)
            except (TypeError, ValueError, ZeroDivisionError):
                pcs = ""    

            allocation_data.append([
                '', '', '',                    # A-C
                i + 1,                         # D - SR.NO
                job.length or "",              # E - Alloca.-1
                count,                         # F - Count (calculated)
                '', '', '', '', '', '', '', '',  # G-L
                consumed_length,              # M - Consu length
                matching_length,              # N - Length (from inventory)
                # '74',                          # O - Pcs (can be dynamic if needed)
                pcs,
                cutting_loss,                 # P - C.L.
                job.total_quantity or ""      # Q - QTY
            ])
        # print('allocation_data:',allocation_data)    

        # Write allocation data with spacing
        for row_data in allocation_data:
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [4,5, 6, 13, 14, 15, 16, 17, 18, 19]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center')
            current_row += 2  # Spacing

        return current_row

        

    def _create_embossing_section(self, ws, start_row,section_jobs):
        """Create the embossing and weight summary section with dynamic data"""

        inventory_data = InventoryItem.objects.select_related('item_id').values_list(
        "id", "item_id__name", "item_id__length"
        )
    

        for i, job in enumerate(section_jobs[:2]):
            try:
                job_length = float(job.length)
            except (TypeError, ValueError):
                job_length = None

            # Find matching inventory length by product name == job.section
            matching_length = None
            for _, prod_name, inv_length in inventory_data:
                if prod_name == job.section:
                    try:
                        matching_length = float(inv_length)
                    except (TypeError, ValueError):
                        matching_length = None
                    break

        thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

        # Embossing header
        ws.merge_cells(f'A{start_row}:M{start_row}')
        ws[f'A{start_row}'] = 'EMBOSSING DETAILS'
        ws[f'A{start_row}'].font = Font(bold=True)
        ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
        ws[f'A{start_row}'].border = thin_border

        # Embossing data headers
        headers = ['EMBOSSING', 'SIN NO', 'SECTION', '', 'Length', 'pcs', 'Unit wt', 'wt in kgs', '', '', 'C.L', 'Pcs', 'Wt in kgs']
        for col_num, value in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col_num, value=value)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Generate embossing data from section_jobs
        embossing_data = []
        for job in section_jobs:
            try:
                unit_wt = float(job.wt_and_pcs) if job.wt_and_pcs else 0
                total_qty = float(job.total_quantity) if job.total_quantity else 0
                wt_in_kgs = round(unit_wt * total_qty, 2)
            except (ValueError, TypeError):
                unit_wt = 0
                wt_in_kgs = 0

            embossing_data.append([
                'MD PG MS21',  # Embossing prefix (can be customized)
                'SIN-36 21-22',  # SIN NO (can be customized)
                job.section or "",  # SECTION from job
                "",  # Empty
                matching_length or "",  # Length from job
                job.total_quantity or "",  # pcs from total_quantity
                unit_wt,  # Unit weight from wt_and_pcs
                wt_in_kgs,  # Calculated weight in kgs
                "", "",  # Empty columns
                "",  # C.L (can be calculated if needed)
                "",  # Pcs (can be duplicated from total_quantity)
                "Kgs"  # Unit
            ])

        # Write embossing data
        for row_num, row_data in enumerate(embossing_data, start_row+2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num in [5, 6, 7, 8, 12]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center')

        # Calculate weight summary from the actual data
        total_planning_wt = sum(row_data[7] for row_data in embossing_data if isinstance(row_data[7], (int, float)))
        total_raw_material_wt = total_planning_wt * 1.01  # Example: adding 1% for raw material
        wastage_percent = 1.05  # Can be calculated or from settings
        wastage_wt = round(total_planning_wt * (wastage_percent / 100), 2)

        # Weight summary - moved to columns O, R, S
        summary_start = start_row + len(embossing_data) -1
        summary = [
            ('Planning Wt.                      :', total_planning_wt, 'Kgs'),
            ('Raw Material Wt.                  :', round(total_raw_material_wt, 2), 'Kgs'),
            ('SRN Wt                            :', '0', 'Kgs'),
            ('Wastage                           :', wastage_percent, '%'),
            ('Wastage Wt.                       :', wastage_wt, 'Kgs')
        ]

        for i, (label, value, unit) in enumerate(summary):
            ws[f'O{summary_start + i}'] = label
            ws[f'O{summary_start + i}'].font = Font(bold=True)
            ws[f'O{summary_start + i}'].alignment = Alignment(horizontal='right')
            
            ws[f'R{summary_start + i}'] = value
            ws[f'R{summary_start + i}'].alignment = Alignment(horizontal='right')
            
            ws[f'S{summary_start + i}'] = unit
            ws[f'S{summary_start + i}'].alignment = Alignment(horizontal='left')



            total_pcs = sum(_safe_int(job.total_quantity) for job in section_jobs)
            total_wt = sum(_safe_float(row[6]) for row in embossing_data if isinstance(row[6], (int, float)))

            # Add total row (matches image format)
            total_row = start_row + len(embossing_data) + 4
            ws.merge_cells(f'C{total_row}')
            ws[f'C{total_row}'] = 'Total:'
            ws[f'C{total_row}'].font = Font(bold=True)
            ws[f'C{total_row}'].alignment = Alignment(horizontal='center')
            
            ws[f'F{total_row}'] = total_pcs
            ws[f'F{total_row}'].font = Font(bold=True)
            ws[f'F{total_row}'].alignment = Alignment(horizontal='center')
            
            ws[f'H{total_row}'] = round(total_wt, 2)
            ws[f'H{total_row}'].font = Font(bold=True)
            ws[f'H{total_row}'].alignment = Alignment(horizontal='center')

            # Add JC number in column I (matches image)
            ws.merge_cells(f'O{total_row}:S{total_row}')
            ws[f'O{total_row}'] = f"JC / 162 / {section_jobs[0].Job_order_no.split('-')[-1]}" if section_jobs and section_jobs[0].Job_order_no else 'JC / 162 / 000'
            ws[f'O{total_row}'].font = Font(bold=True)
            ws[f'O{total_row}'].alignment = Alignment(horizontal='center')

        return summary_start + len(summary) + 1

        
        
    def _check_inventory_availability(self, job):
        """Check if inventory is available for this job and return detailed status"""
        try:
            inventory_item = InventoryItem.objects.filter(item_id__name=job.section).first()
            if not inventory_item:
                return {
                    'available': False,
                    'reason': f"No inventory found for section: {job.section}"
                }
            
            try:
                job_length = float(job.length)
                inv_length = float(inventory_item.item_id.length)
                if job_length > inv_length:
                    return {
                        'available': False,
                        'reason': f"Job length ({job_length}mm) exceeds inventory length ({inv_length}mm)"
                    }
            except (TypeError, ValueError) as e:
                return {
                    'available': False,
                    'reason': f"Invalid length values: {str(e)}"
                }
                
            return {
                'available': True,
                'reason': "Inventory available"
            }
        except Exception as e:
            return {
                'available': False,
                'reason': f"Error checking inventory: {str(e)}"
            }     
        
        
        

    def job_card_creation(self, request):
        try:
            project_id = request.GET.get('project_id')
            if not project_id:
                return Response({"error": "project_id not provided"}, status=status.HTTP_400_BAD_REQUEST)
    
            try:
                project = Project.objects.get(id=project_id)
            except Project.DoesNotExist:
                return Response({"error": "Project not found"}, status=status.HTTP_404_NOT_FOUND)
    
            if not project.bom_sheet:
                return Response({"error": "No BOM Sheet linked to this project"}, status=status.HTTP_404_NOT_FOUND)
    
            bom_sheet = project.bom_sheet
            job_cards = JobCard.objects.filter(bom_sheet=bom_sheet).order_by('section')
    
            all_jobs_created = True
            jc_creation_status = []
            jc_create_info = {}
    
            wb = Workbook()
            ws = wb.active
            ws.title = "Job_Cards"
    
            column_widths = {
                'A': 15, 'B': 10, 'C': 28, 'D': 15, 'E': 10, 'F': 10, 'G': 10,
                'H': 10, 'I': 5, 'J': 5, 'K': 18, 'L': 10, 'M': 10, 'N': 2,
                'O': 28, 'P': 8, 'Q': 5, 'R': 5, 'S': 15
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
    
            current_row = 2
            section_groups = defaultdict(list)
    
            for job in job_cards:
                section_groups[job.section].append(job)
                inventory_status = self._check_inventory_availability(job)
                jc_created = inventory_status['available']
                jc_create_info[job.id] = jc_created
    
                jc_creation_status.append({
                    'job_id': job.id,
                    'section': job.section,
                    'length': job.length,
                    'created': jc_created,
                    'remarks': inventory_status['reason']
                })
    
                if not jc_created:
                    all_jobs_created = False
    
            # ? Update BOM sheet with jc_create status
            try:
                bom_file_path = bom_sheet.excel_file.path
                bom_wb = load_workbook(bom_file_path)
                bom_ws = bom_wb.active  # do NOT rename or overwrite this sheet
            
                # Detect columns
                header_map = {}
                for col in range(1, bom_ws.max_column + 1):
                    col_name = bom_ws.cell(row=1, column=col).value
                    if col_name:
                        header_map[col_name.strip().lower()] = col
            
                # Add jc_create column if not present
                if 'jc_create' not in header_map:
                    jc_create_col = bom_ws.max_column + 1
                    bom_ws.cell(row=1, column=jc_create_col, value='jc_create')
                else:
                    jc_create_col = header_map['jc_create']
            
                # Ensure job ID column exists
                if 'job id' in header_map:
                    job_id_col = header_map['job id']
                else:
                    raise Exception("Job ID column not found in BOM sheet. Please ensure there's a 'Job ID' column header.")
            
                # Update each row
                for row in range(2, bom_ws.max_row + 1):
                    job_id_cell = bom_ws.cell(row=row, column=job_id_col).value
                    if job_id_cell in jc_create_info:
                        bom_ws.cell(row=row, column=jc_create_col, value=bool(jc_create_info[job_id_cell]))
            
                # Save changes to the same BOM file
                bom_wb.save(bom_file_path)
            
            except Exception as e:
                print(f"Error updating BOM sheet with jc_create status: {str(e)}")
    
            # ? Update BOM sheet status in DB
            bom_sheet.jc_produced = all_jobs_created
            bom_sheet.last_processed_at = timezone.now()
            bom_sheet.save()
    
            # ? Generate Job Cards
            for section, jobs in section_groups.items():
                start_job_card_row = current_row
                valid_jobs = [job for job in jobs if jc_create_info.get(job.id, False)]
                if not valid_jobs:
                    continue
    
                current_row = self._create_job_card_header(ws, current_row, valid_jobs[0])
                current_row = self.generate_section(ws, current_row, valid_jobs)
                current_row = self._create_allocation_section(ws, current_row, valid_jobs)
                current_row = self._create_embossing_section(ws, current_row, valid_jobs)
                self._add_outside_border(ws, start_job_card_row, current_row - 1)
                current_row += 2
    
            # ? JC Creation Status Sheet
            status_sheet = wb.create_sheet("JC_Creation_Status")
            headers = ["Job ID", "Section", "Length (mm)", "JC Created", "Remarks"]
            status_sheet.append(headers)
    
            for col in range(1, len(headers) + 1):
                cell = status_sheet.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
    
            for status in jc_creation_status:
                row = [
                    status['job_id'],
                    status['section'],
                    status['length'],
                    "YES" if status['created'] else "NO",
                    status['remarks']
                ]
                status_sheet.append(row)
    
            for col in ['A', 'B', 'C', 'D', 'E']:
                status_sheet.column_dimensions[col].width = 20
    
            # ? Save Excel file and return download response
            save_dir = os.path.join(settings.MEDIA_ROOT, "job_cards")
            os.makedirs(save_dir, exist_ok=True)
    
            filename = f"job_cards_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            full_path = os.path.join(save_dir, filename)
            wb.save(full_path)
    
            with open(full_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
    
        except Exception as e:
            if 'bom_sheet' in locals():
                bom_sheet.jc_produced = False
                bom_sheet.save()
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)




    





   
    
    
    
    
    
    
    
    
    
class JobCardViewSet(viewsets.ModelViewSet):


    serializer_class = GetJobCardSerializer 
    
    
    def _generate_next_jc_no(self):
        last_job_card = JobCard.objects.order_by('-id').first()
        if last_job_card and last_job_card.jc_no:
            try:
                # Extract the numeric part from jc_no, e.g., JC-005 ? 5
                last_number = int(last_job_card.jc_no.split('-')[-1])
            except ValueError:
                last_number = 0
        else:
            last_number = 0

        next_number = last_number + 1
        return f"JC-{next_number:03d}"



    def post(self, request):
        try:
            # For POST requests, get data from request.data
            project_id = request.GET.get("project_id")
            print('project_id:', project_id)
            if not project_id:
                return JsonResponse({"error": "Project ID is required"}, status=400)

            project = Project.objects.get(id=project_id)
            
            if not project.bom_sheet or not project.bom_sheet.bom_file:
                return JsonResponse({"error": "No BOM file associated with this project."}, status=400)

            # Process the BOM file
            bom_file = project.bom_sheet.bom_file
            with bom_file.open("rb") as f:
                df = pd.read_excel(f)

            # Clean and validate data
            df.rename(columns=lambda x: x.strip(), inplace=True)
            df = df.astype(str).applymap(lambda x: x.strip() if isinstance(x, str) else x)

            # Validate required columns
            required_columns = {
                "Vrno": "Job Order No",
                "Tower Type Name": "Tower Type",
                "Markno": "Mark No",
                "Product Name": "Section",
                "Length": "Length",
                "Width": "Width",
                "Section Weight": "Unit",
                "Wt/Piece": "WT/PCS",
                "Quantity": "Total Quantity",
                "Blk.Wt.(Kg)": "Total Weight",
                "Remark": "Remarks"
            }

            missing_cols = [col for col in required_columns.keys() if col not in df.columns]
            if missing_cols:
                return JsonResponse({
                    "error": f"Missing required columns: {', '.join(missing_cols)}",
                    "required_columns": list(required_columns.keys())
                }, status=400)

            created_job_cards = []
            inventory_data = self._get_inventory_data()

            for _, row in df.iterrows():
                # Skip rows without job order number
                job_order_no = str(row.get("Vrno", "")).strip()
                if not job_order_no:
                    continue

                # Create JobCard with basic info
                job_card = JobCard(
                    jc_no=self._generate_next_jc_no(),
                    bom_sheet=project.bom_sheet,
                    project=project,
                    Job_order_no=job_order_no,
                    type_of_tower=str(row.get("Tower Type Name", "")).strip(),
                    mark_no=str(row.get("Markno", "")).strip(),
                    section=str(row.get("Product Name", "")).strip(),
                    length=str(row.get("Length", "")).strip(),
                    width=str(row.get("Width", "")).strip(),
                    unit=str(row.get("Section Weight", "")).strip(),
                    wt_and_pcs=str(row.get("Wt/Piece", "")).strip(),
                    total_quantity=str(row.get("Quantity", "")).strip(),
                    total_weight=str(row.get("Blk.Wt.(Kg)", "")).strip(),
                    remarks=str(row.get("Remark", "")).strip(),
                    product_code=str(row.get("Product", "")).strip() if "Product" in df.columns else "",
                    tower_type=str(row.get("Tower Type Name", "")).strip(),
                    client_name=str(row.get("Party Name", "")).strip() if "Party Name" in df.columns else ""
                )

                # Process allocation data
                self._process_allocation_data(job_card, inventory_data)
                
                # Save the job card
                job_card.save()
                created_job_cards.append(job_card)

                # Create JobCardReport with calculations
                self._create_job_card_report(job_card)

            if not created_job_cards:
                return JsonResponse({
                    "success": False,
                    "message": "No valid job cards were created from the BOM file."
                }, status=400)

            # Serialize and return response
            serializer = JobCardSerializer(created_job_cards, many=True)
            return JsonResponse({
                "success": True,
                "count": len(created_job_cards),
                "data": serializer.data
            }, status=201)

        except Project.DoesNotExist:
            return JsonResponse({"error": "Project not found."}, status=404)
        except Exception as e:
            return JsonResponse({
                "error": f"Failed to process BOM file: {str(e)}"
            }, status=500)

    def _get_inventory_data(self):
        """Helper method to get inventory data for calculations"""
        return InventoryItem.objects.select_related('item_id').values(
            'id',
            'item_name',
            'item_code',
            'quantity_available_pc',
            'item_id__length',
            'item_id__name',
            'item_id__product_no'
        )

    def _process_allocation_data(self, job_card, inventory_data):
        """Process and save allocation data to job card"""
        try:
            job_length = float(job_card.length) if job_card.length else None
            total_quantity = float(job_card.total_quantity) if job_card.total_quantity else None
    
            best_match = None
            section_name = job_card.section
    
            # Find best matching inventory item
            for item in inventory_data:
                if (str(item['item_name']).upper() == section_name.upper() or 
                    str(item['item_id__name']).upper() == section_name.upper()):
                    try:
                        inv_length = float(item['item_id__length']) if item['item_id__length'] else None
                        avail_qty = float(item['quantity_available_pc']) if item['quantity_available_pc'] else 0
    
                        if inv_length and avail_qty > 0:
                            count = int(inv_length // job_length) if job_length else 0
                            efficiency = (count * job_length) / inv_length if inv_length else 0
                            if not best_match or efficiency > best_match['efficiency']:
                                best_match = {
                                    "id": item['id'],
                                    "item_code": item['item_code'],
                                    "length": inv_length,
                                    "available_quantity": avail_qty,
                                    "efficiency": efficiency,
                                    "product_no": item['item_id__product_no']
                                }
                    except (TypeError, ValueError):
                        continue
    
            if best_match:
                try:
                    job_length_decimal = Decimal(str(job_length))
                    inventory_length_decimal = Decimal(str(best_match['length']))
                    total_quantity_decimal = Decimal(str(total_quantity))
            
                    count = int(inventory_length_decimal // job_length_decimal) if job_length_decimal > 0 else 0
                    consumed_length = count * job_length_decimal
                    cutting_loss = inventory_length_decimal - consumed_length
                    pieces_needed = (total_quantity_decimal / Decimal(count)) if count > 0 else Decimal(0)
            
                    print(">>> Matching Inventory Found:")
                    print("  - count:", count)
                    print("  - consumed_length:", float(consumed_length))
                    print("  - cutting_loss:", float(cutting_loss))
                    print("  - pieces:", float(pieces_needed))
            
                    # Update inventory quantity
                    inventory_item = InventoryItem.objects.get(id=best_match["id"])
                    inventory_item.quantity_available_pc = max(Decimal(0), inventory_item.quantity_available_pc - pieces_needed)
                    inventory_item.save()
            
                    # Update job card
                    job_card.allocated_1 = str(job_length_decimal)
                    job_card.count = str(count)
                    job_card.consumed_length = str(consumed_length)
                    job_card.inventory_length = str(inventory_length_decimal)
                    job_card.cutting_loss = str(cutting_loss)
                    job_card.serial_number = best_match['item_code']
                    job_card.available_length = str(inventory_length_decimal)
                    job_card.pieces = str(pieces_needed)
            
                    job_card.save()
            
                except (InvalidOperation, ZeroDivisionError, Exception) as e:
                    print("Error processing allocation data:", e)
    
        except Exception as e:
            print(f"Error processing allocation data: {str(e)}")
            
            

    def _create_job_card_report(self, job_card):
        """Create and save JobCardReport with calculations"""
        try:
            # Calculate weight data
            try:
                unit_wt = float(job_card.wt_and_pcs) if job_card.wt_and_pcs else 0
                total_qty = float(job_card.total_quantity) if job_card.total_quantity else 0
                calculated_weight = round(unit_wt * total_qty, 2)
            except (ValueError, TypeError):
                calculated_weight = 0

            # Calculate summary weights
            total_planning_wt = calculated_weight
            total_raw_material_wt = total_planning_wt * 1.01  # Adding 1% for raw material
            wastage_percent = 1.05
            wastage_wt = round(total_planning_wt * (wastage_percent / 100), 2)

            # Create the report
            JobCardReport.objects.create(
                job_card=job_card,
                planning_weight=total_planning_wt,
                raw_material_weight=round(total_raw_material_wt, 2),
                srn_weight=0,
                wastage_percent=wastage_percent,
                wastage_weight=wastage_wt,
                calculated_weight=calculated_weight
            )
            
            job_card.pcs = str(int(total_qty))  # cast to string since pcs is a CharField
            job_card.unit_weight = unit_wt
            job_card.weight_in_kgs = calculated_weight
            job_card.save()

        except Exception as e:
            print(f"Error creating job card report: {str(e)}")
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
class GetJobCardViewSet(viewsets.ModelViewSet):


    queryset = JobCard.objects.all()
    serializer_class = GetJobCardSerializer  # This serializer should handle nested fields like `reports`

    def retrieve_job_by_id(self, request):

        job_card_id = request.GET.get('id')

        if not job_card_id:
            return Response({"success": False, "error": "ID parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            job_card = JobCard.objects.get(id=job_card_id)
            serializer = self.get_serializer(job_card)
            return Response({"success": True, "data": serializer.data}, status=status.HTTP_200_OK)
        except JobCard.DoesNotExist:
            return Response({"success": False, "error": "Job Card not found."}, status=status.HTTP_404_NOT_FOUND)    
    
    
    def get_list_of_job_card(self, request):
    
        project_id = request.GET.get('project_id')
        if not project_id:
            return Response({"error": "project_id is required"}, status=status.HTTP_400_BAD_REQUEST)
    
        try:
            project = Project.objects.get(id=project_id)
        except Project.DoesNotExist:
            return Response({"error": "Project not found"}, status=status.HTTP_404_NOT_FOUND)
    
        if not project.bom_sheet:
            return Response({"error": "No BOM Sheet associated with this project"}, status=status.HTTP_404_NOT_FOUND)
    
        job_cards = JobCard.objects.filter(bom_sheet=project.bom_sheet)
        serializer = JobCardMinimalSerializer(job_cards, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)  
    
      








   






















        
        
        
    